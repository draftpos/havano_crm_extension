from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging
import re

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _strip_quotes(val):
    """Remove surrounding double-quotes that ERPNext adds to IDs."""
    if isinstance(val, str):
        val = val.strip()
        if val.startswith('"') and val.endswith('"'):
            val = val[1:-1]
    return val


def _strip_html(val):
    """Strip HTML tags from rich-text fields."""
    if isinstance(val, str):
        val = re.sub(r'<[^>]+>', ' ', val)
        val = re.sub(r'\s+', ' ', val).strip()
    return val


def _safe_str(val, default=''):
    """Convert a value to string safely."""
    if val is None:
        return default
    s = _strip_quotes(str(val)).strip()
    return s if s and s.lower() != 'none' else default


def _clean_datetime(val):
    """
    Clean ERPNext datetime strings for Odoo.
    ERPNext format: '2024-07-16 14:25:10.834878' (microseconds included)
    Odoo expects:   '2024-07-16 14:25:10'
    """
    if val is None:
        return False
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    s = str(val).strip()
    if not s or s.lower() == 'none':
        return False
    s = s.replace('T', ' ')
    if '.' in s:
        s = s[:s.index('.')]
    if len(s) == 10:  # If it's just a date, append time
        s += ' 00:00:00'
    return s if len(s) >= 10 else False


def _clean_date(val):
    """
    Clean date strings (e.g. '03-11-2025' or '2025-11-03') to Odoo YYYY-MM-DD format.
    """
    if val is None:
        return False
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.lower() == 'none':
        return False
    if ' ' in s:
        s = s.split(' ')[0]
    parts = re.split(r'[-/]', s)
    if len(parts) == 3:
        p1, p2, p3 = parts[0], parts[1], parts[2]
        if len(p1) == 4:  # YYYY-MM-DD
            return f"{p1}-{p2.zfill(2)}-{p3.zfill(2)}"
        elif len(p3) == 4:  # DD-MM-YYYY or MM-DD-YYYY
            return f"{p3}-{p2.zfill(2)}-{p1.zfill(2)}"
    return False




def _parse_erpnext_xlsx(file_bytes):
    """
    Parse an ERPNext Data Import Template (.xlsx).

    Structure:
      Row 1:  'Data Import Template'
      Row 2:  'Table:  <DocType>'
      ...
      Row 16: 'Column Name:' | field1 | field2 | ...  <- HEADERS
      Row 17: 'Mandatory:'
      Row 18: 'Type:'
      Row 19: 'Info:'
      Row 20: 'Start entering data below'
      Row 21+: NULL | value1 | value2 | ...            <- DATA (col 0 always null)

    Returns: (headers: list[str], records: list[dict])
    """
    import zipfile
    
    # First pass: find headers
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    except zipfile.BadZipFile:
        raise UserError(_("The uploaded file is not a valid Excel (.xlsx) file."))
    except Exception as e:
        raise UserError(_("Could not open Excel file: %s") % str(e))

    ws = wb.active
    headers = None
    data_start_row = None
    is_template = False
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
        first = str(row[0]).strip() if row[0] else ""
        if first == "Column Name:":
            headers = [str(c).strip() if c is not None else "" for c in row[1:]]
            data_start_row = i + 1 + 5
            is_template = True
            break
    wb.close()

    if headers is None:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
        headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        data_start_row = 2
        wb.close()

    records = []
    col_offset = 1 if is_template else 0

    # Second pass: read data
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for idx, h in enumerate(headers):
            if h:
                data_col = idx + col_offset
                row_dict[h] = row[data_col] if data_col < len(row) else None
        records.append(row_dict)
    wb.close()
    
    return headers, records


class LegacyDataImportWizard(models.TransientModel):
    _name = 'legacy.data.import.wizard'
    _description = 'Wizard to import legacy Leads and To-Dos from ERPNext'

    import_type = fields.Selection([
        ('lead', 'Leads (Lead.xlsx)'),
        ('opportunity', 'Opportunities (Opportunity.xlsx)'),
        ('todo', 'To-Dos (ToDo.xlsx)'),
        ('contact', 'Contacts (Contact.xlsx)'),
        ('item', 'Items/Products (Item.xlsx)'),
        ('user', 'Users (User.xlsx)'),
    ], string='What are you importing?', required=True, default='lead')

    data_file = fields.Binary(string='Excel File (.xlsx)', required=True)
    filename = fields.Char(string='Filename')

    state = fields.Selection([
        ('upload', 'Upload'),
        ('importing', 'Importing'),
        ('done', 'Done')
    ], string='Status', default='upload')
    
    total_count = fields.Integer(string='Total Records', default=0)
    imported_count = fields.Integer(string='Processed Records', default=0)
    progress = fields.Float(string='Progress', compute='_compute_progress')

    @api.depends('imported_count', 'total_count')
    def _compute_progress(self):
        for rec in self:
            if rec.total_count > 0:
                rec.progress = (rec.imported_count / rec.total_count) * 100.0
            else:
                rec.progress = 0.0

    def action_import_data(self):
        if not openpyxl:
            raise UserError(_("Python library 'openpyxl' is required. Please install it."))
        if not self.data_file:
            raise UserError(_("Please upload a file first."))

        file_bytes = base64.b64decode(self.data_file)
        headers, records = _parse_erpnext_xlsx(file_bytes)

        if not records:
            raise UserError(_("No data rows found in the uploaded file. Please check the file format."))

        if self.state == 'upload':
            self.total_count = len(records)
            self.imported_count = 0
            self.state = 'importing'

        batch_size = 500
        start_idx = self.imported_count
        end_idx = start_idx + batch_size
        batch_records = records[start_idx:end_idx]

        if not batch_records:
            self.state = 'done'
            return self._return_success_notification()

        # We will only process this batch
        records = batch_records

        created = 0
        skipped = 0

        if self.import_type == 'lead':
            leads_to_create = []
            for r in records:
                name = (
                    _safe_str(r.get('first_name')) or
                    _safe_str(r.get('lead_name')) or
                    _safe_str(r.get('company_name')) or
                    'Unknown Lead'
                )
                # Deal size — stored as float in ERPNext
                deal_size = 0.0
                try:
                    deal_size = float(r.get('custom_deal_size_') or r.get('custom_deal_size') or 0)
                except (TypeError, ValueError):
                    pass

                # Demo done flag — ERPNext stores as 0/1
                demo_raw = r.get('custom_demo_done_') or r.get('custom_demo_done')
                demo_done = bool(demo_raw and str(demo_raw).strip() not in ('0', '', 'None', 'False'))

                # Quote status — map to our selection
                quote_raw = _safe_str(r.get('custom_quote'))
                quote_val = quote_raw if quote_raw in ('Quote', 'Not yet Quoted') else False

                vals = {
                    'name': name,
                    'partner_name': _safe_str(r.get('company_name')),
                    'phone': _safe_str(r.get('mobile_no')),
                    'email_from': _safe_str(r.get('email_id')),
                    # ERPNext ID
                    'custom_naming_series':    _safe_str(r.get('name')),
                    'custom_lead_source':      _safe_str(r.get('source')),
                    'custom_territory':        _safe_str(r.get('territory')),
                    'custom_industry':         _safe_str(r.get('industry')),
                    # Extra ERPNext fields
                    'custom_product':          _safe_str(r.get('custom_product')),
                    'custom_deal_size':        deal_size,
                    'custom_type_of_business': _safe_str(r.get('custom_type_of_business')),
                    'custom_quote':            quote_val,
                    'custom_quote_date':       _clean_date(r.get('custom_quote_date')),
                    'custom_technician':       _safe_str(r.get('custom_technician')),
                    'custom_demo_done':        demo_done,
                }
                
                # Map Salesperson (owner)
                owner_str = _safe_str(r.get('owner'))
                if owner_str:
                    user = self.env['res.users'].search([
                        '|', ('login', '=', owner_str), ('name', 'ilike', owner_str)
                    ], limit=1)
                    if user:
                        vals['user_id'] = user.id

                creation = _clean_datetime(r.get('creation'))
                if creation:
                    vals['legacy_create_date'] = creation
                    vals['create_date'] = creation
                leads_to_create.append(vals)

            if leads_to_create:
                batch_size = 200
                for i in range(0, len(leads_to_create), batch_size):
                    batch = leads_to_create[i:i + batch_size]
                    self.env['crm.lead'].create(batch)
                    self.env.cr.commit()
                created = len(leads_to_create)

        elif self.import_type == 'opportunity':
            leads_to_create = []
            for r in records:
                name = (
                    _safe_str(r.get('party_name')) or
                    _safe_str(r.get('customer_name')) or
                    _safe_str(r.get('name')) or
                    'Unknown Opportunity'
                )
                vals = {
                    'name': name,
                    'partner_name': name,
                    'custom_naming_series': _safe_str(r.get('name')),
                    'custom_lead_source': _safe_str(r.get('source')),
                    'probability': float(r.get('probability') or 0),
                    'type': 'opportunity',
                }
                
                # Map Salesperson (owner)
                owner_str = _safe_str(r.get('owner'))
                if owner_str:
                    user = self.env['res.users'].search([
                        '|', ('login', '=', owner_str), ('name', 'ilike', owner_str)
                    ], limit=1)
                    if user:
                        vals['user_id'] = user.id

                creation = _clean_datetime(r.get('creation')) or _clean_datetime(r.get('transaction_date'))
                if creation:
                    vals['legacy_create_date'] = creation
                    vals['create_date'] = creation
                leads_to_create.append(vals)

            if leads_to_create:
                self.env['crm.lead'].create(leads_to_create)
                created = len(leads_to_create)

        elif self.import_type == 'todo':
            todos_to_create = []
            for r in records:
                # reference_name is the ERPNext Lead ID (e.g. "CRM-LEAD-2024-00003")
                erpnext_lead_id = _safe_str(r.get('reference_name'))
                lead = False
                if erpnext_lead_id:
                    # Try to find the Odoo lead by matching the stored ERPNext ID
                    lead = self.env['crm.lead'].search(
                        [('custom_naming_series', '=', erpnext_lead_id)], limit=1
                    )

                description = _strip_html(_safe_str(r.get('description')))
                subject = description[:80] if description else 'Imported ToDo'

                vals = {
                    'name': subject,
                    'status': _safe_str(r.get('status'), 'Open'),
                    'allocated_to': _safe_str(r.get('owner')),
                    'lead_id': lead.id if lead else False,
                }

                date_val = _clean_date(r.get('date'))
                if date_val:
                    vals['date'] = date_val

                creation = _clean_datetime(r.get('creation'))
                if creation:
                    vals['legacy_create_date'] = creation
                    vals['create_date'] = creation

                todos_to_create.append(vals)

            if todos_to_create:
                batch_size = 200
                for i in range(0, len(todos_to_create), batch_size):
                    batch = todos_to_create[i:i + batch_size]
                    self.env['todo.task'].create(batch)
                    self.env.cr.commit()
                created = len(todos_to_create)

        elif self.import_type == 'contact':
            partners_to_create = []
            for r in records:
                name = (
                    _safe_str(r.get('full_name')) or
                    _safe_str(r.get('lead_name')) or
                    _safe_str(r.get('customer_name')) or
                    _safe_str(r.get('name')) or
                    'Unknown Contact'
                )
                vals = {
                    'name': name,
                    'phone': _safe_str(r.get('mobile_no') or r.get('phone')),
                    'email': _safe_str(r.get('email_id')),
                    'is_company': False,
                    'customer_rank': 1,
                }
                partners_to_create.append(vals)

            if partners_to_create:
                self.env['res.partner'].create(partners_to_create)
                created = len(partners_to_create)

        elif self.import_type == 'item':
            products_to_create = []
            for r in records:
                name = (
                    _safe_str(r.get('item_name')) or
                    _safe_str(r.get('item_code')) or
                    _safe_str(r.get('name')) or
                    'Unknown Item'
                )
                vals = {
                    'name': name,
                    'default_code': _safe_str(r.get('item_code') or r.get('name')),
                    'type': 'consu',
                    'sale_ok': True,
                    'purchase_ok': True,
                }
                categ_name = _safe_str(r.get('item_group'))
                if categ_name:
                    categ = self.env['product.category'].search([('name', '=', categ_name)], limit=1)
                    if categ:
                        vals['categ_id'] = categ.id
                products_to_create.append(vals)

            if products_to_create:
                self.env['product.template'].create(products_to_create)
                created = len(products_to_create)

        elif self.import_type == 'user':
            for r in records:
                email = _safe_str(r.get('email') or r.get('email_id'))
                name  = _safe_str(r.get('full_name') or r.get('first_name') or email)
                if not email:
                    skipped += 1
                    continue
                existing = self.env['res.users'].sudo().search(
                    [('login', '=', email)], limit=1
                )
                if existing:
                    skipped += 1
                    continue
                try:
                    self.env['res.users'].sudo().with_context(
                        no_reset_password=True,
                        mail_create_nolog=True,
                        mail_notrack=True,
                    ).create({
                        'name':      name or email,
                        'login':     email,
                        'email':     email,
                        'password':  'HavanoImport@2024',
                        'groups_id': [(4, self.env.ref('base.group_user').id)],
                    })
                    created += 1
                except Exception as e:
                    _logger.warning("Skipped user %s: %s", email, str(e))
                    skipped += 1

        self.imported_count += len(records)

        if self.imported_count >= self.total_count:
            self.state = 'done'
            return self._return_success_notification()

        # Return action to reload the wizard for the next batch
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'legacy.data.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _return_success_notification(self):
        msg = _('Import complete! %s records processed.') % self.imported_count
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Complete ✓'),
                'message': msg,
                'type': 'success',
                'sticky': True,
            }
        }
